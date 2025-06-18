import re
import pandas as pd
from bs4 import BeautifulSoup
from typing import List, Dict, Any, Optional
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import time

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


class WebsiteAnalyzer:
    """Web sitesi analizörü - HTML kaynağını analiz eder ve tıklanabilir elementleri tespit eder"""
    
    def __init__(self):
        self.clickable_elements = []
        self.analysis_results = {}
    
    def analyze_html_source(self, html_source: str, url: str = "", analysis_type: str = "quick") -> Dict[str, Any]:
        """
        HTML kaynağını analiz eder ve tıklanabilir elementleri tespit eder
        
        Args:
            html_source: Analiz edilecek HTML kaynağı
            url: Web sitesi URL'si (opsiyonel)
            analysis_type: Analiz türü ("quick", "detailed", "full")
            
        Returns:
            Analiz sonuçları sözlüğü
        """
        try:
            soup = BeautifulSoup(html_source, 'html.parser')
            
            # Tıklanabilir elementleri tespit et
            clickable_elements = self._find_clickable_elements(soup)
            
            # Form elementlerini tespit et
            form_elements = self._find_form_elements(soup)
            
            # Navigasyon elementlerini tespit et
            navigation_elements = self._find_navigation_elements(soup)
            
            # Analiz türüne göre ek analizler
            script_analysis = {}
            page_analysis = {}
            
            if analysis_type in ["detailed", "full"]:
                # Script ve style analizi
                script_analysis = self._analyze_scripts_and_styles(soup)
                
                # Genel sayfa analizi
                page_analysis = self._analyze_page_structure(soup)
            
            # Analiz sonuçlarını hazırla
            analysis_results = {
                "url": url,
                "analysis_date": datetime.now().isoformat(),
                "analysis_type": analysis_type,
                "total_clickable_elements": len(clickable_elements),
                "total_form_elements": len(form_elements),
                "total_navigation_elements": len(navigation_elements),
                "clickable_elements": clickable_elements,
                "form_elements": form_elements,
                "navigation_elements": navigation_elements,
                "script_analysis": script_analysis,
                "page_analysis": page_analysis,
                "summary": self._generate_summary(clickable_elements, form_elements, navigation_elements, script_analysis, page_analysis)
            }
            
            self.analysis_results = analysis_results
            return analysis_results
            
        except Exception as e:
            return {
                "error": f"HTML analizi sırasında hata oluştu: {str(e)}",
                "url": url,
                "analysis_date": datetime.now().isoformat(),
                "analysis_type": analysis_type
            }
    
    def analyze_dynamic_website(self, url: str, wait_time: int = 10, login_credentials: Dict[str, str] = None) -> Dict[str, Any]:
        """
        Dinamik web sitesi analizi (JavaScript ile render edilen siteler için)
        
        Args:
            url: Analiz edilecek web sitesi URL'si
            wait_time: Sayfa yüklenmesi için beklenecek süre (saniye)
            login_credentials: Giriş bilgileri (username, password, username_selector, password_selector, submit_selector)
            
        Returns:
            Analiz sonuçları sözlüğü
        """
        # URL validation
        if not url:
            return {"error": "URL boş olamaz"}
        
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        if not SELENIUM_AVAILABLE:
            return {
                "error": "Selenium kütüphanesi bulunamadı. Dinamik analiz için 'pip install selenium' komutunu çalıştırın."
            }
        
        driver = None
        try:
            # Chrome options
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # Görünmez modda çalıştır
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            # WebDriver başlat
            driver = webdriver.Chrome(options=chrome_options)
            driver.set_page_load_timeout(30)
            driver.implicitly_wait(10)
            
            # Sayfayı yükle
            print(f"Loading URL: {url}")
            driver.get(url)
            
            # Sayfanın yüklenmesini bekle
            time.sleep(2)
            
            # Sayfanın tamamen yüklenmesini bekle
            try:
                WebDriverWait(driver, 10).until(
                    lambda driver: driver.execute_script("return document.readyState") == "complete"
                )
            except:
                print("Page load timeout, continuing anyway...")
            
            # Check if page is accessible (not blocked)
            page_title = driver.title.lower()
            page_source = driver.page_source.lower()
            
            # Check for common blocking indicators
            blocking_indicators = [
                "access denied", "blocked", "forbidden", "captcha", "verify you are human",
                "cloudflare", "ddos protection", "rate limit", "too many requests",
                "erişim engellendi", "erişim reddedildi", "doğrulama gerekli"
            ]
            
            for indicator in blocking_indicators:
                if indicator in page_title or indicator in page_source:
                    driver.quit()
                    return {
                        "error": f"Web sitesi erişimi engellendi. '{indicator}' tespit edildi.\n\n" +
                                "Çözüm önerileri:\n" +
                                "1. Birkaç dakika bekleyip tekrar deneyin\n" +
                                "2. Farklı bir URL deneyin\n" +
                                "3. Web sitesi bot koruması kullanıyor olabilir\n" +
                                "4. Statik HTML analizi kullanmayı deneyin"
                    }
            
            # Login işlemi (eğer credentials verilmişse)
            if login_credentials and login_credentials.get('username') and login_credentials.get('password'):
                # Check if login form exists
                if not self._detect_login_form(driver):
                    driver.quit()
                    return {
                        "error": "Bu sayfada login formu bulunamadı. Lütfen şunları kontrol edin:\n" +
                                "1. URL doğru mu?\n" +
                                "2. Sayfa tamamen yüklendi mi?\n" +
                                "3. Login formu JavaScript ile mi yükleniyor?\n" +
                                "4. Sayfa login gerektiriyor mu?"
                    }
                
                login_success = self._perform_login(driver, login_credentials)
                if not login_success:
                    driver.quit()
                    return {
                        "error": "Giriş yapılamadı. Lütfen şunları kontrol edin:\n" +
                                "1. Kullanıcı adı ve şifre doğru mu?\n" +
                                "2. Web sitesi login formu var mı?\n" +
                                "3. Özel selector'lar doğru mu?\n" +
                                "4. Web sitesi erişilebilir mi?\n" +
                                "5. Captcha veya 2FA var mı?"
                    }
            
            # Sayfanın yüklenmesini bekle
            time.sleep(wait_time)
            
            # Sayfa kaynağını al
            page_source = driver.page_source
            
            # HTML analizi yap
            analysis_results = self.analyze_html_source(page_source, url)
            
            # Dinamik element analizi ekle
            dynamic_analysis = self._analyze_dynamic_elements(driver)
            analysis_results.update(dynamic_analysis)
            
            # Sonuçları sakla
            self.analysis_results = analysis_results
            
            return analysis_results
            
        except Exception as e:
            print(f"Dynamic analysis error for URL {url}: {str(e)}")
            return {
                "error": f"Dinamik analiz sırasında hata oluştu: {str(e)}\n\n" +
                        "Olası nedenler:\n" +
                        "1. Web sitesi erişilebilir değil\n" +
                        "2. Web sitesi çok yavaş yükleniyor\n" +
                        "3. Web sitesi bot erişimini engelliyor\n" +
                        "4. ChromeDriver bulunamadı\n" +
                        "5. Sistem kaynakları yetersiz"
            }
        finally:
            if driver:
                driver.quit()
    
    def _detect_login_form(self, driver) -> bool:
        """Sayfada login formu olup olmadığını tespit eder"""
        try:
            # Common login form indicators
            login_indicators = [
                'input[name="username"]',
                'input[name="email"]',
                'input[type="email"]',
                'input[name="password"]',
                'input[type="password"]',
                'form[action*="login"]',
                'form[action*="signin"]',
                '.login-form',
                '#login-form',
                'form[class*="login"]'
            ]
            
            for selector in login_indicators:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        return True
                except:
                    continue
            
            return False
            
        except Exception as e:
            print(f"Login form detection error: {e}")
            return False

    def _perform_login(self, driver, credentials: Dict[str, str]) -> bool:
        """
        Web sitesinde giriş yapar
        
        Args:
            driver: Selenium WebDriver
            credentials: Giriş bilgileri
            
        Returns:
            Giriş başarılı mı?
        """
        try:
            username = credentials.get('username')
            password = credentials.get('password')
            
            # Default selectors to try
            default_username_selectors = [
                'input[name="username"]',
                'input[name="email"]', 
                'input[type="email"]',
                '#username',
                '#email',
                'input[placeholder*="user"]',
                'input[placeholder*="email"]',
                'input[placeholder*="User"]',
                'input[placeholder*="Email"]'
            ]
            
            default_password_selectors = [
                'input[name="password"]',
                'input[type="password"]',
                '#password',
                'input[placeholder*="password"]',
                'input[placeholder*="Password"]'
            ]
            
            default_submit_selectors = [
                'button[type="submit"]',
                'input[type="submit"]',
                '.login-btn',
                '#login-btn',
                'button[class*="login"]',
                'button[class*="submit"]',
                'input[value*="Login"]',
                'input[value*="Sign In"]',
                'input[value*="Giriş"]',
                'button[class*="btn"]',
                'input[class*="btn"]'
            ]
            
            # Use custom selectors if provided, otherwise use defaults
            username_selectors = [credentials.get('username_selector')] if credentials.get('username_selector') else default_username_selectors
            password_selectors = [credentials.get('password_selector')] if credentials.get('password_selector') else default_password_selectors
            submit_selectors = [credentials.get('submit_selector')] if credentials.get('submit_selector') else default_submit_selectors
            
            # Find username field
            username_field = None
            for selector in username_selectors:
                try:
                    username_field = driver.find_element(By.CSS_SELECTOR, selector)
                    if username_field.is_displayed():
                        break
                except:
                    continue
            
            if not username_field:
                print(f"Username field not found with selectors: {username_selectors}")
                return False
            
            # Find password field
            password_field = None
            for selector in password_selectors:
                try:
                    password_field = driver.find_element(By.CSS_SELECTOR, selector)
                    if password_field.is_displayed():
                        break
                except:
                    continue
            
            if not password_field:
                print(f"Password field not found with selectors: {password_selectors}")
                return False
            
            # Find submit button
            submit_button = None
            for selector in submit_selectors:
                try:
                    submit_button = driver.find_element(By.CSS_SELECTOR, selector)
                    if submit_button.is_displayed():
                        break
                except:
                    continue
            
            # If CSS selectors fail, try XPath for text-based button finding
            if not submit_button:
                xpath_selectors = [
                    "//button[contains(text(), 'Login')]",
                    "//button[contains(text(), 'Sign In')]",
                    "//button[contains(text(), 'Giriş')]",
                    "//button[contains(text(), 'Submit')]",
                    "//input[@type='submit']",
                    "//button[@type='submit']"
                ]
                
                for xpath in xpath_selectors:
                    try:
                        submit_button = driver.find_element(By.XPATH, xpath)
                        if submit_button.is_displayed():
                            break
                    except:
                        continue
            
            if not submit_button:
                print(f"Submit button not found with selectors: {submit_selectors}")
                return False
            
            # Fill in credentials
            username_field.clear()
            username_field.send_keys(username)
            
            password_field.clear()
            password_field.send_keys(password)
            
            # Click submit button
            submit_button.click()
            
            # Wait for login to complete
            time.sleep(3)
            
            return True
            
        except Exception as e:
            print(f"Login hatası: {str(e)}")
            return False
    
    def _scroll_page(self, driver):
        """Sayfayı aşağı kaydırarak lazy loading elementlerini yükler"""
        try:
            # Sayfanın sonuna kadar scroll et
            last_height = driver.execute_script("return document.body.scrollHeight")
            
            while True:
                # Sayfanın sonuna scroll et
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                
                # Yeni içeriğin yüklenmesini bekle
                time.sleep(2)
                
                # Yeni yüksekliği hesapla
                new_height = driver.execute_script("return document.body.scrollHeight")
                
                # Eğer yükseklik değişmediyse, tüm içerik yüklenmiş demektir
                if new_height == last_height:
                    break
                    
                last_height = new_height
                
        except Exception as e:
            print(f"Scroll sırasında hata: {e}")
    
    def _find_dynamic_elements(self, driver) -> Dict[str, Any]:
        """Dinamik olarak yüklenen elementleri tespit eder"""
        dynamic_elements = {
            "clickable_elements": [],
            "form_elements": [],
            "navigation_elements": []
        }
        
        try:
            # Tüm tıklanabilir elementleri bul
            clickable_selectors = [
                "button", "a", "[onclick]", "[role='button']", 
                "[tabindex]", ".btn", ".button", ".clickable"
            ]
            
            for selector in clickable_selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    try:
                        if element.is_displayed():
                            element_info = {
                                "tag": element.tag_name,
                                "text": element.text[:100] if element.text else "",
                                "class": element.get_attribute("class") or "",
                                "id": element.get_attribute("id") or "",
                                "href": element.get_attribute("href") or "",
                                "onclick": element.get_attribute("onclick") or "",
                                "role": element.get_attribute("role") or ""
                            }
                            dynamic_elements["clickable_elements"].append(element_info)
                    except:
                        continue
            
            # Form elementlerini bul
            form_selectors = ["input", "select", "textarea", "form"]
            for selector in form_selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    try:
                        if element.is_displayed():
                            element_info = {
                                "tag": element.tag_name,
                                "type": element.get_attribute("type") or "",
                                "name": element.get_attribute("name") or "",
                                "id": element.get_attribute("id") or "",
                                "placeholder": element.get_attribute("placeholder") or "",
                                "class": element.get_attribute("class") or ""
                            }
                            dynamic_elements["form_elements"].append(element_info)
                    except:
                        continue
            
            # Navigasyon elementlerini bul
            nav_selectors = ["nav", "[role='navigation']", ".nav", ".menu", ".navbar"]
            for selector in nav_selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    try:
                        if element.is_displayed():
                            element_info = {
                                "tag": element.tag_name,
                                "text": element.text[:100] if element.text else "",
                                "class": element.get_attribute("class") or "",
                                "id": element.get_attribute("id") or ""
                            }
                            dynamic_elements["navigation_elements"].append(element_info)
                    except:
                        continue
                        
        except Exception as e:
            print(f"Dinamik element tespiti sırasında hata: {e}")
        
        return dynamic_elements
    
    def _detect_javascript_events(self, driver) -> Dict[str, Any]:
        """JavaScript event listener'larını tespit eder"""
        js_events = {
            "onclick_events": [],
            "onchange_events": [],
            "onsubmit_events": [],
            "other_events": []
        }
        
        try:
            # onclick event'lerini bul
            onclick_elements = driver.find_elements(By.CSS_SELECTOR, "[onclick]")
            for element in onclick_elements:
                js_events["onclick_events"].append({
                    "tag": element.tag_name,
                    "text": element.text[:50] if element.text else "",
                    "onclick": element.get_attribute("onclick")
                })
            
            # onchange event'lerini bul
            onchange_elements = driver.find_elements(By.CSS_SELECTOR, "[onchange]")
            for element in onchange_elements:
                js_events["onchange_events"].append({
                    "tag": element.tag_name,
                    "type": element.get_attribute("type") or "",
                    "onchange": element.get_attribute("onchange")
                })
            
            # onsubmit event'lerini bul
            onsubmit_elements = driver.find_elements(By.CSS_SELECTOR, "[onsubmit]")
            for element in onsubmit_elements:
                js_events["onsubmit_events"].append({
                    "tag": element.tag_name,
                    "onsubmit": element.get_attribute("onsubmit")
                })
                
        except Exception as e:
            print(f"JavaScript event tespiti sırasında hata: {e}")
        
        return js_events
    
    def _find_clickable_elements(self, soup: BeautifulSoup) -> List[Dict[str, Any]]:
        """Tıklanabilir elementleri tespit eder"""
        clickable_elements = []
        
        # Butonlar
        buttons = soup.find_all(['button', 'input'], type=['button', 'submit', 'reset'])
        for button in buttons:
            element_info = self._extract_element_info(button, "Button")
            if element_info:
                clickable_elements.append(element_info)
        
        # Linkler
        links = soup.find_all('a', href=True)
        for link in links:
            element_info = self._extract_element_info(link, "Link")
            if element_info:
                clickable_elements.append(element_info)
        
        # JavaScript ile tıklanabilir elementler
        js_clickable = soup.find_all(attrs={'onclick': True})
        for element in js_clickable:
            element_info = self._extract_element_info(element, "JavaScript Clickable")
            if element_info:
                clickable_elements.append(element_info)
        
        # CSS ile tıklanabilir elementler (cursor: pointer)
        css_clickable = soup.find_all(attrs={'style': re.compile(r'cursor:\s*pointer', re.I)})
        for element in css_clickable:
            element_info = self._extract_element_info(element, "CSS Clickable")
            if element_info:
                clickable_elements.append(element_info)
        
        # Div'ler ve span'lar (genellikle tıklanabilir)
        divs_spans = soup.find_all(['div', 'span'], class_=re.compile(r'(btn|button|click|link)', re.I))
        for element in divs_spans:
            element_info = self._extract_element_info(element, "Div/Span Clickable")
            if element_info:
                clickable_elements.append(element_info)
        
        return clickable_elements
    
    def _find_form_elements(self, soup: BeautifulSoup) -> List[Dict[str, Any]]:
        """Form elementlerini tespit eder"""
        form_elements = []
        
        # Form alanları
        form_inputs = soup.find_all('input')
        for input_elem in form_inputs:
            element_info = self._extract_element_info(input_elem, "Form Input")
            if element_info:
                form_elements.append(element_info)
        
        # Select elementleri
        selects = soup.find_all('select')
        for select in selects:
            element_info = self._extract_element_info(select, "Select Dropdown")
            if element_info:
                form_elements.append(element_info)
        
        # Textarea elementleri
        textareas = soup.find_all('textarea')
        for textarea in textareas:
            element_info = self._extract_element_info(textarea, "Textarea")
            if element_info:
                form_elements.append(element_info)
        
        return form_elements
    
    def _find_navigation_elements(self, soup: BeautifulSoup) -> List[Dict[str, Any]]:
        """Navigasyon elementlerini tespit eder"""
        navigation_elements = []
        
        # Menü elementleri
        menu_elements = soup.find_all(['nav', 'ul', 'ol'], class_=re.compile(r'(menu|nav|navigation)', re.I))
        for menu in menu_elements:
            links = menu.find_all('a')
            for link in links:
                element_info = self._extract_element_info(link, "Navigation Link")
                if element_info:
                    navigation_elements.append(element_info)
        
        # Breadcrumb elementleri
        breadcrumbs = soup.find_all(class_=re.compile(r'breadcrumb', re.I))
        for breadcrumb in breadcrumbs:
            links = breadcrumb.find_all('a')
            for link in links:
                element_info = self._extract_element_info(link, "Breadcrumb Link")
                if element_info:
                    navigation_elements.append(element_info)
        
        return navigation_elements
    
    def _extract_element_info(self, element, element_type: str) -> Optional[Dict[str, Any]]:
        """Element bilgilerini çıkarır"""
        try:
            # Element metni
            text = element.get_text(strip=True)
            if not text and element.get('placeholder'):
                text = element.get('placeholder')
            if not text and element.get('alt'):
                text = element.get('alt')
            if not text and element.get('title'):
                text = element.get('title')
            
            # Element ID'si
            element_id = element.get('id', '')
            
            # CSS sınıfları
            classes = ' '.join(element.get('class', []))
            
            # XPath benzeri yol
            xpath = self._generate_xpath(element)
            
            # CSS seçici
            css_selector = self._generate_css_selector(element)
            
            # Href (linkler için)
            href = element.get('href', '')
            
            # Type (input'lar için)
            input_type = element.get('type', '')
            
            # Name (form elementleri için)
            name = element.get('name', '')
            
            # Priority (önem derecesi)
            priority = self._determine_priority(element, element_type)
            
            return {
                "type": element_type,
                "text": text[:100] if text else "No text",  # Maksimum 100 karakter
                "id": element_id,
                "classes": classes,
                "xpath": xpath,
                "css_selector": css_selector,
                "href": href,
                "input_type": input_type,
                "name": name,
                "priority": priority,
                "element_tag": element.name,
                "is_visible": self._is_element_visible(element),
                "has_text": bool(text),
                "recommended_action": self._get_recommended_action(element_type, text, href)
            }
            
        except Exception as e:
            print(f"Element bilgisi çıkarılırken hata: {e}")
            return None
    
    def _generate_xpath(self, element) -> str:
        """Element için XPath benzeri yol oluşturur"""
        try:
            path_parts = []
            current = element
            
            while current and current.name:
                tag = current.name
                
                # ID varsa kullan
                if current.get('id'):
                    path_parts.append(f"//{tag}[@id='{current.get('id')}']")
                    break
                
                # Class varsa kullan
                if current.get('class'):
                    classes = '.'.join(current.get('class'))
                    path_parts.append(f"//{tag}[contains(@class,'{classes}')]")
                    break
                
                # Pozisyon kullan
                siblings = current.find_previous_siblings(tag)
                position = len(siblings) + 1
                path_parts.append(f"//{tag}[{position}]")
                
                current = current.parent
            
            return ''.join(reversed(path_parts)) if path_parts else f"//{element.name}"
            
        except Exception:
            return f"//{element.name}"
    
    def _generate_css_selector(self, element) -> str:
        """Element için CSS seçici oluşturur"""
        try:
            if element.get('id'):
                return f"#{element.get('id')}"
            
            if element.get('class'):
                classes = '.'.join(element.get('class'))
                return f"{element.name}.{classes}"
            
            return element.name
            
        except Exception:
            return element.name
    
    def _determine_priority(self, element, element_type: str) -> str:
        """Element önceliğini belirler"""
        # Yüksek öncelikli elementler
        high_priority = ['submit', 'login', 'register', 'buy', 'purchase', 'checkout', 'add to cart']
        # Orta öncelikli elementler
        medium_priority = ['menu', 'nav', 'search', 'filter', 'sort']
        
        text = element.get_text(strip=True).lower()
        href = element.get('href', '').lower()
        
        for keyword in high_priority:
            if keyword in text or keyword in href:
                return "High"
        
        for keyword in medium_priority:
            if keyword in text or keyword in href:
                return "Medium"
        
        return "Low"
    
    def _is_element_visible(self, element) -> bool:
        """Element görünür olup olmadığını kontrol eder"""
        # Basit görünürlük kontrolü
        style = element.get('style', '')
        if 'display: none' in style or 'visibility: hidden' in style:
            return False
        
        # Hidden attribute kontrolü
        if element.get('hidden'):
            return False
        
        return True
    
    def _get_recommended_action(self, element_type: str, text: str, href: str) -> str:
        """Önerilen test aksiyonunu belirler"""
        text_lower = text.lower()
        href_lower = href.lower()
        
        if element_type == "Button":
            if any(keyword in text_lower for keyword in ['submit', 'save', 'confirm']):
                return "Click and verify form submission"
            elif any(keyword in text_lower for keyword in ['delete', 'remove', 'cancel']):
                return "Click and verify confirmation dialog"
            else:
                return "Click and verify page change or action"
        
        elif element_type == "Link":
            if href_lower.startswith('http'):
                return "Click and verify external link opens"
            elif href_lower.startswith('#'):
                return "Click and verify page scroll or anchor"
            else:
                return "Click and verify navigation to new page"
        
        elif element_type == "Form Input":
            return "Enter test data and verify validation"
        
        elif element_type == "Select Dropdown":
            return "Select option and verify value change"
        
        else:
            return "Click and verify expected behavior"
    
    def _generate_summary(self, clickable_elements, form_elements, navigation_elements, script_analysis, page_analysis) -> Dict[str, Any]:
        """Analiz özeti oluşturur"""
        return {
            "total_elements": len(clickable_elements) + len(form_elements) + len(navigation_elements),
            "element_types": {
                "buttons": len([e for e in clickable_elements if e["type"] == "Button"]),
                "links": len([e for e in clickable_elements if e["type"] == "Link"]),
                "form_inputs": len([e for e in form_elements if e["type"] == "Form Input"]),
                "dropdowns": len([e for e in form_elements if e["type"] == "Select Dropdown"]),
                "navigation": len(navigation_elements)
            },
            "priority_distribution": {
                "high": len([e for e in clickable_elements if e["priority"] == "High"]),
                "medium": len([e for e in clickable_elements if e["priority"] == "Medium"]),
                "low": len([e for e in clickable_elements if e["priority"] == "Low"])
            },
            "script_analysis": script_analysis,
            "page_analysis": page_analysis
        }
    
    def generate_excel_checklist(self, output_path: str = None) -> str:
        """
        Analiz sonuçlarını Excel checklist formatında oluşturur
        
        Args:
            output_path: Çıktı dosyası yolu (opsiyonel)
            
        Returns:
            Oluşturulan dosyanın yolu
        """
        if not self.analysis_results:
            raise ValueError("Önce HTML analizi yapılmalıdır")
        
        # Excel dosyası oluştur
        wb = Workbook()
        
        # Ana sayfa
        ws_main = wb.active
        ws_main.title = "Website Analysis Summary"
        
        # Özet bilgiler
        ws_main['A1'] = "Website Analysis Report"
        ws_main['A1'].font = Font(size=16, bold=True)
        
        ws_main['A3'] = "URL:"
        ws_main['B3'] = self.analysis_results.get('url', 'N/A')
        
        ws_main['A4'] = "Analysis Date:"
        ws_main['B4'] = datetime.fromisoformat(self.analysis_results['analysis_date']).strftime('%Y-%m-%d %H:%M:%S')
        
        ws_main['A6'] = "Total Clickable Elements:"
        ws_main['B6'] = self.analysis_results['total_clickable_elements']
        
        ws_main['A7'] = "Total Form Elements:"
        ws_main['B7'] = self.analysis_results['total_form_elements']
        
        ws_main['A8'] = "Total Navigation Elements:"
        ws_main['B8'] = self.analysis_results['total_navigation_elements']
        
        # Script ve Style analizi
        script_analysis = self.analysis_results.get('script_analysis', {})
        ws_main['A10'] = "Script Analysis:"
        ws_main['A10'].font = Font(bold=True)
        ws_main['A11'] = "Total Scripts:"
        ws_main['B11'] = script_analysis.get('total_scripts', 0)
        ws_main['A12'] = "External Scripts:"
        ws_main['B12'] = script_analysis.get('external_scripts', 0)
        ws_main['A13'] = "Detected Framework:"
        ws_main['B13'] = script_analysis.get('is_spa', 'Unknown')
        
        # Sayfa yapısı analizi
        page_analysis = self.analysis_results.get('page_analysis', {})
        ws_main['A15'] = "Page Structure:"
        ws_main['A15'].font = Font(bold=True)
        ws_main['A16'] = "Total Divs:"
        ws_main['B16'] = page_analysis.get('total_divs', 0)
        ws_main['A17'] = "Total Links:"
        ws_main['B17'] = page_analysis.get('total_links', 0)
        ws_main['A18'] = "Has Loading Indicator:"
        ws_main['B18'] = "Yes" if page_analysis.get('has_loading_indicator', False) else "No"
        ws_main['A19'] = "Is Minimal Content:"
        ws_main['B19'] = "Yes" if page_analysis.get('is_minimal_content', False) else "No"
        
        # Tıklanabilir elementler sayfası
        ws_clickable = wb.create_sheet("Clickable Elements")
        self._create_clickable_elements_sheet(ws_clickable)
        
        # Form elementleri sayfası
        ws_forms = wb.create_sheet("Form Elements")
        self._create_form_elements_sheet(ws_forms)
        
        # Navigasyon elementleri sayfası
        ws_nav = wb.create_sheet("Navigation Elements")
        self._create_navigation_elements_sheet(ws_nav)
        
        # Test checklist sayfası
        ws_checklist = wb.create_sheet("Test Checklist")
        self._create_test_checklist_sheet(ws_checklist)
        
        # Dosyayı kaydet
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"website_analysis_{timestamp}.xlsx"
        
        # Exports klasörü yoksa oluştur
        os.makedirs('exports', exist_ok=True)
        full_path = os.path.join('exports', output_path)
        
        wb.save(full_path)
        return full_path
    
    def _create_clickable_elements_sheet(self, ws):
        """Tıklanabilir elementler sayfasını oluşturur"""
        headers = [
            "Element Type", "Text", "ID", "Classes", "XPath", "CSS Selector", 
            "Href", "Priority", "Visible", "Recommended Action", "Test Status", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for element in self.analysis_results['clickable_elements']:
            ws.cell(row=row, column=1, value=element['type'])
            ws.cell(row=row, column=2, value=element['text'])
            ws.cell(row=row, column=3, value=element['id'])
            ws.cell(row=row, column=4, value=element['classes'])
            ws.cell(row=row, column=5, value=element['xpath'])
            ws.cell(row=row, column=6, value=element['css_selector'])
            ws.cell(row=row, column=7, value=element['href'])
            ws.cell(row=row, column=8, value=element['priority'])
            ws.cell(row=row, column=9, value="Yes" if element['is_visible'] else "No")
            ws.cell(row=row, column=10, value=element['recommended_action'])
            ws.cell(row=row, column=11, value="Not Tested")
            ws.cell(row=row, column=12, value="")
            row += 1
        
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_form_elements_sheet(self, ws):
        """Form elementleri sayfasını oluşturur"""
        headers = [
            "Element Type", "Text", "ID", "Name", "Input Type", "Classes", 
            "XPath", "CSS Selector", "Required", "Recommended Action", "Test Status", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for element in self.analysis_results['form_elements']:
            ws.cell(row=row, column=1, value=element['type'])
            ws.cell(row=row, column=2, value=element['text'])
            ws.cell(row=row, column=3, value=element['id'])
            ws.cell(row=row, column=4, value=element['name'])
            ws.cell(row=row, column=5, value=element['input_type'])
            ws.cell(row=row, column=6, value=element['classes'])
            ws.cell(row=row, column=7, value=element['xpath'])
            ws.cell(row=row, column=8, value=element['css_selector'])
            ws.cell(row=row, column=9, value="Yes" if element.get('required') else "No")
            ws.cell(row=row, column=10, value=element['recommended_action'])
            ws.cell(row=row, column=11, value="Not Tested")
            ws.cell(row=row, column=12, value="")
            row += 1
    
    def _create_navigation_elements_sheet(self, ws):
        """Navigasyon elementleri sayfasını oluşturur"""
        headers = [
            "Element Type", "Text", "ID", "Classes", "Href", "XPath", 
            "CSS Selector", "Priority", "Recommended Action", "Test Status", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(color="000000", bold=True)
        
        row = 2
        for element in self.analysis_results['navigation_elements']:
            ws.cell(row=row, column=1, value=element['type'])
            ws.cell(row=row, column=2, value=element['text'])
            ws.cell(row=row, column=3, value=element['id'])
            ws.cell(row=row, column=4, value=element['classes'])
            ws.cell(row=row, column=5, value=element['href'])
            ws.cell(row=row, column=6, value=element['xpath'])
            ws.cell(row=row, column=7, value=element['css_selector'])
            ws.cell(row=row, column=8, value=element['priority'])
            ws.cell(row=row, column=9, value=element['recommended_action'])
            ws.cell(row=row, column=10, value="Not Tested")
            ws.cell(row=row, column=11, value="")
            row += 1
    
    def _create_test_checklist_sheet(self, ws):
        """Test checklist sayfasını oluşturur"""
        headers = [
            "Test ID", "Element Type", "Element Text", "Priority", "Test Action", 
            "Expected Result", "Test Status", "Tester", "Date", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        test_id = 1
        row = 2
        
        # Tüm elementleri birleştir
        all_elements = (
            self.analysis_results['clickable_elements'] +
            self.analysis_results['form_elements'] +
            self.analysis_results['navigation_elements']
        )
        
        # Önceliğe göre sırala
        all_elements.sort(key=lambda x: {"High": 3, "Medium": 2, "Low": 1}.get(x.get('priority', 'Low'), 1), reverse=True)
        
        for element in all_elements:
            ws.cell(row=row, column=1, value=f"TC-{test_id:03d}")
            ws.cell(row=row, column=2, value=element['type'])
            ws.cell(row=row, column=3, value=element['text'])
            ws.cell(row=row, column=4, value=element['priority'])
            ws.cell(row=row, column=5, value=element['recommended_action'])
            ws.cell(row=row, column=6, value="Element should respond correctly")
            ws.cell(row=row, column=7, value="Not Started")
            ws.cell(row=row, column=8, value="")
            ws.cell(row=row, column=9, value="")
            ws.cell(row=row, column=10, value="")
            row += 1
            test_id += 1
        
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _analyze_scripts_and_styles(self, soup: BeautifulSoup) -> Dict[str, Any]:
        """Script ve style dosyalarını analiz eder"""
        scripts = soup.find_all('script')
        styles = soup.find_all('link', rel='stylesheet')
        
        script_sources = []
        for script in scripts:
            src = script.get('src', '')
            if src:
                script_sources.append(src)
        
        style_sources = []
        for style in styles:
            href = style.get('href', '')
            if href:
                style_sources.append(href)
        
        return {
            "total_scripts": len(scripts),
            "external_scripts": len(script_sources),
            "total_styles": len(styles),
            "external_styles": len(style_sources),
            "script_sources": script_sources[:5],  # İlk 5 script
            "style_sources": style_sources[:5],   # İlk 5 style
            "is_spa": self._detect_spa_framework(script_sources),
            "has_async_scripts": any(script.get('async') for script in scripts),
            "has_defer_scripts": any(script.get('defer') for script in scripts)
        }
    
    def _analyze_page_structure(self, soup: BeautifulSoup) -> Dict[str, Any]:
        """Sayfa yapısını analiz eder"""
        # Ana elementleri say
        divs = soup.find_all('div')
        spans = soup.find_all('span')
        buttons = soup.find_all('button')
        links = soup.find_all('a')
        forms = soup.find_all('form')
        inputs = soup.find_all('input')
        
        # Loading spinner kontrolü
        loading_indicators = soup.find_all(text=re.compile(r'loading|spinner|progress', re.I))
        loading_elements = soup.find_all(class_=re.compile(r'loading|spinner|progress', re.I))
        
        # Meta tag analizi
        meta_tags = soup.find_all('meta')
        title = soup.find('title')
        
        return {
            "total_divs": len(divs),
            "total_spans": len(spans),
            "total_buttons": len(buttons),
            "total_links": len(links),
            "total_forms": len(forms),
            "total_inputs": len(inputs),
            "has_loading_indicator": len(loading_indicators) > 0 or len(loading_elements) > 0,
            "loading_elements_count": len(loading_elements),
            "meta_tags_count": len(meta_tags),
            "has_title": title is not None,
            "title_text": title.get_text() if title else None,
            "is_minimal_content": len(divs) < 5 and len(links) < 3 and len(buttons) < 2
        }
    
    def _detect_spa_framework(self, script_sources: List[str]) -> str:
        """SPA framework'ünü tespit eder"""
        for src in script_sources:
            if 'react' in src.lower():
                return "React"
            elif 'vue' in src.lower():
                return "Vue"
            elif 'angular' in src.lower():
                return "Angular"
            elif 'next' in src.lower():
                return "Next.js"
            elif 'nuxt' in src.lower():
                return "Nuxt.js"
        
        return "Unknown/Static"
    
    def _analyze_dynamic_elements(self, driver) -> Dict[str, Any]:
        """
        Dinamik elementleri analiz eder
        
        Args:
            driver: Selenium WebDriver
            
        Returns:
            Dinamik analiz sonuçları
        """
        try:
            # Sayfanın scroll edilmesi (lazy loading için)
            self._scroll_page(driver)
            
            # Dinamik elementleri tespit et
            dynamic_elements = self._find_dynamic_elements(driver)
            
            # JavaScript event listener'ları tespit et
            js_events = self._detect_javascript_events(driver)
            
            return {
                "dynamic_elements": dynamic_elements,
                "javascript_events": js_events,
                "page_title": driver.title,
                "current_url": driver.current_url,
                "window_size": driver.get_window_size()
            }
            
        except Exception as e:
            return {
                "dynamic_elements": [],
                "javascript_events": [],
                "error": f"Dinamik analiz hatası: {str(e)}"
            } 