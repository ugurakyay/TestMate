import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

class ExcelProcessor:
    """Excel dosyalarını işlemek için sınıf"""
    
    def __init__(self):
        self.template_columns = [
            'Test ID',
            'Test Adı',
            'Test Açıklaması',
            'Öncelik',
            'Test Türü',
            'Adım No',
            'Adım Açıklaması',
            'Locator Türü',
            'Locator Değeri',
            'Eylem',
            'Test Verisi',
            'Beklenen Sonuç',
            'Durum',
            'Notlar'
        ]
    
    def create_test_scenario_template(self, filename: str = "test_scenarios_template.xlsx") -> str:
        """Test senaryoları için Excel şablonu oluştur"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Senaryoları"
        
        # Başlık satırını ekle
        for col, header in enumerate(self.template_columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Örnek veri ekle
        example_data = [
            ["TC001", "Kullanıcı Girişi", "Kullanıcının email ve şifre ile giriş yapabilmesi", "Yüksek", "Web", 1, "Tarayıcıyı aç", "", "", "Aç", "https://example.com", "Sayfa yüklendi", "", ""],
            ["TC002", "Email Girişi", "Email alanına geçerli email adresi girilmesi", "Yüksek", "Web", 1, "Email alanına tıkla", "id", "email-input", "Tıkla", "", "Email alanı aktif", "", ""],
            ["TC003", "Şifre Girişi", "Şifre alanına geçerli şifre girilmesi", "Yüksek", "Web", 1, "Şifre alanına tıkla", "id", "password-input", "Tıkla", "", "Şifre alanı aktif", "", ""],
            ["TC004", "Giriş Butonu", "Giriş butonuna tıklanarak giriş yapılması", "Yüksek", "Web", 1, "Giriş butonuna tıkla", "id", "login-btn", "Tıkla", "", "Giriş yapıldı", "", ""],
            ["TC005", "Hatalı Email", "Geçersiz email ile giriş denemesi", "Orta", "Web", 1, "Email alanına geçersiz email gir", "id", "email-input", "Yaz", "invalid@email", "Hata mesajı görüntülendi", "", ""],
            ["TC006", "Hatalı Şifre", "Geçersiz şifre ile giriş denemesi", "Orta", "Web", 1, "Şifre alanına geçersiz şifre gir", "id", "password-input", "Yaz", "wrongpass", "Hata mesajı görüntülendi", "", ""],
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Sütun genişliklerini ayarla
        column_widths = [8, 20, 30, 10, 10, 8, 25, 12, 20, 10, 20, 20, 10, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Açıklama sayfası ekle
        self._add_instructions_sheet(wb)
        
        # Dosyayı kaydet
        filepath = os.path.join("exports", filename)
        os.makedirs("exports", exist_ok=True)
        wb.save(filepath)
        return filepath
    
    def _add_instructions_sheet(self, wb: Workbook):
        """Açıklama sayfası ekle"""
        ws = wb.create_sheet("Kullanım Talimatları")
        
        instructions = [
            ["Alan", "Açıklama", "Örnek"],
            ["Test ID", "Benzersiz test kimliği", "TC001"],
            ["Test Adı", "Test senaryosunun adı", "Kullanıcı Girişi"],
            ["Test Açıklaması", "Test senaryosunun detaylı açıklaması", "Kullanıcının email ve şifre ile giriş yapabilmesi"],
            ["Öncelik", "Test önceliği (Yüksek/Orta/Düşük)", "Yüksek"],
            ["Test Türü", "Test türü (Web/Mobil/API)", "Web"],
            ["Adım No", "Test adımının sıra numarası", "1"],
            ["Adım Açıklaması", "Adımın açıklaması", "Tarayıcıyı aç"],
            ["Locator Türü", "Element bulma yöntemi (id/class/xpath/css)", "id"],
            ["Locator Değeri", "Element bulma değeri", "email-input"],
            ["Eylem", "Yapılacak eylem (Tıkla/Yaz/Bekle/Seç)", "Tıkla"],
            ["Test Verisi", "Girilecek veri", "test@example.com"],
            ["Beklenen Sonuç", "Beklenen sonuç", "Email alanı aktif"],
            ["Durum", "Test durumu (Boş bırakılabilir)", ""],
            ["Notlar", "Ek notlar", ""]
        ]
        
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
    
    def read_test_scenarios(self, filepath: str) -> List[Dict[str, Any]]:
        """Excel dosyasından test senaryolarını oku"""
        try:
            df = pd.read_excel(filepath, sheet_name="Test Senaryoları")
            
            # NaN değerleri temizle
            df = df.fillna('')
            
            # Veriyi test senaryolarına dönüştür
            test_scenarios = []
            
            for index, row in df.iterrows():
                test_id = row.get('Test ID', '')
                if pd.isna(test_id) or test_id == '':
                    continue
                
                # Her satırı ayrı bir test senaryosu olarak işle
                # Eğer aynı Test ID varsa, adım numarasına göre benzersiz yap
                unique_test_id = str(test_id)
                if any(scenario['test_id'] == unique_test_id for scenario in test_scenarios):
                    step_number = row.get('Adım No', 1)
                    unique_test_id = f"{test_id}_step_{step_number}"
                
                test_scenario = {
                    'test_id': unique_test_id,
                    'test_name': str(row.get('Test Adı', f'Test {unique_test_id}')),
                    'test_description': str(row.get('Test Açıklaması', '')),
                    'priority': str(row.get('Öncelik', 'Orta')),
                    'test_type': str(row.get('Test Türü', 'Web')),
                    'steps': []
                }
                
                # Tek adımlı test senaryosu oluştur
                step = {
                    'step_number': int(row.get('Adım No', 1)) if pd.notna(row.get('Adım No')) else 1,
                    'description': str(row.get('Adım Açıklaması', '')),
                    'locator_type': str(row.get('Locator Türü', '')),
                    'locator_value': str(row.get('Locator Değeri', '')),
                    'action': str(row.get('Eylem', '')),
                    'test_data': str(row.get('Test Verisi', '')),
                    'expected_result': str(row.get('Beklenen Sonuç', '')),
                    'status': str(row.get('Durum', '')),
                    'notes': str(row.get('Notlar', ''))
                }
                
                test_scenario['steps'].append(step)
                test_scenarios.append(test_scenario)
            
            return test_scenarios
            
        except Exception as e:
            raise Exception(f"Excel dosyası okuma hatası: {str(e)}")
    
    def export_test_results(self, results: List[Dict[str, Any]], filename: str = None) -> str:
        """Test sonuçlarını Excel'e export et"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_results_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sonuçları"
        
        # Başlık satırı
        headers = [
            'Test ID', 'Test Adı', 'Test Türü', 'Durum', 'Başlangıç Zamanı', 
            'Bitiş Zamanı', 'Süre (sn)', 'Hata Mesajı', 'Ekran Görüntüsü'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Sonuçları ekle
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result.get('test_id', ''))
            ws.cell(row=row_idx, column=2, value=result.get('test_name', ''))
            ws.cell(row=row_idx, column=3, value=result.get('test_type', ''))
            
            status_cell = ws.cell(row=row_idx, column=4, value=result.get('status', ''))
            if result.get('status') == 'PASSED':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif result.get('status') == 'FAILED':
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            ws.cell(row=row_idx, column=5, value=result.get('start_time', ''))
            ws.cell(row=row_idx, column=6, value=result.get('end_time', ''))
            ws.cell(row=row_idx, column=7, value=result.get('duration', ''))
            ws.cell(row=row_idx, column=8, value=result.get('error_message', ''))
            ws.cell(row=row_idx, column=9, value=result.get('screenshot_path', ''))
        
        # Sütun genişliklerini ayarla
        column_widths = [10, 25, 10, 10, 20, 20, 10, 40, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Dosyayı kaydet
        filepath = os.path.join("exports", filename)
        os.makedirs("exports", exist_ok=True)
        wb.save(filepath)
        return filepath
    
    def validate_test_scenarios(self, scenarios: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Test senaryolarını doğrula"""
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'summary': {
                'total_tests': len(scenarios),
                'total_steps': 0,
                'tests_with_errors': 0,
                'tests_with_warnings': 0
            }
        }
        
        for scenario in scenarios:
            test_errors = []
            test_warnings = []
            
            # Zorunlu alanları kontrol et
            if not scenario.get('test_id'):
                test_errors.append("Test ID eksik")
            
            if not scenario.get('test_name'):
                test_errors.append("Test adı eksik")
            
            if not scenario.get('steps'):
                test_errors.append("Test adımları eksik")
            
            # Test adımlarını kontrol et
            for step in scenario.get('steps', []):
                validation_result['summary']['total_steps'] += 1
                
                if not step.get('description'):
                    test_warnings.append(f"Adım {step.get('step_number', '?')}: Açıklama eksik")
                
                if step.get('action') in ['Tıkla', 'Yaz'] and not step.get('locator_value'):
                    test_warnings.append(f"Adım {step.get('step_number', '?')}: Eylem için locator eksik")
                
                if step.get('action') == 'Yaz' and not step.get('test_data'):
                    test_warnings.append(f"Adım {step.get('step_number', '?')}: Yazma eylemi için test verisi eksik")
            
            if test_errors:
                validation_result['errors'].extend([f"{scenario.get('test_id', 'Bilinmeyen')}: {error}" for error in test_errors])
                validation_result['summary']['tests_with_errors'] += 1
                validation_result['valid'] = False
            
            if test_warnings:
                validation_result['warnings'].extend([f"{scenario.get('test_id', 'Bilinmeyen')}: {warning}" for warning in test_warnings])
                validation_result['summary']['tests_with_warnings'] += 1
        
        return validation_result 